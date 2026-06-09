from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

'''
FULL REVIEW
113 failures
repository-wide grounding analysis

INPUT_CSV = Path(
    "results/runs/scidata_4293/analysis/grounding_failure_taxonomy.csv"
)

OUTPUT_XLSX = Path(
    "results/runs/scidata_4293/analysis/grounding_failure_review_v2.xlsx"
)
'''

'''
TABULAR REVIEW
39 unresolved + 9 ambiguous
semantic validation paper analysis
'''
INPUT_CSV = Path(
    "results/runs/scidata_4293/analysis/grounding_failure_taxonomy_tabular_only.csv"
)

OUTPUT_XLSX = Path(
    "results/runs/scidata_4293/analysis/grounding_failure_review_tabular.xlsx"
)

CATEGORY_OPTIONS = [
    "external_artifact",
    "missing_software_artifact",
    "format_substitution",
    "repository_missing",
    "unsupported_modality",
    "inventory_absent",
    "schema_granularity_mismatch",
    "semantic_ambiguity",
]


SUPPORT_OPTIONS = [
    "unsupported",
    "partial_support",
    "supported",
    "not_applicable",
]


CATEGORY_DEFINITIONS = [
    {
        "category": "external_artifact",
        "中文解释": "资源存在于当前下载仓库之外",
        "definition": (
            "The referenced artifact is external to the downloaded repository "
            "boundary, such as Zenodo, Figshare, API endpoint, web tool, or "
            "remote database."
        ),
        "example": "Zenodo Archive; API Data Stream; Interactive Visualization Tool",
    },
    {
        "category": "missing_software_artifact",
        "中文解释": "文档提到的软件、脚本或 notebook 未发布",
        "definition": (
            "The documentation references code or software assets, but the "
            "downloaded repository does not contain corresponding scripts, "
            "notebooks, query files, or executables."
        ),
        "example": "No .py, .R, .ipynb, .rq, or executable file found",
    },
    {
        "category": "format_substitution",
        "中文解释": "语义对应物可能存在，但物理格式不同",
        "definition": (
            "A semantically related artifact appears to exist, but in a "
            "different format from the logical claim, such as CSV instead of "
            "NetCDF or XLSX instead of CSV."
        ),
        "example": "Expected NetCDF, but similar data appears in CSV",
    },
    {
        "category": "repository_missing",
        "中文解释": "文档声明的资源在仓库中确实找不到",
        "definition": (
            "The documentation claims a file or artifact, but no physical file "
            "or plausible substitute is present in the downloaded repository."
        ),
        "example": "No file named codebook.csv and no plausible data dictionary",
    },
    {
        "category": "unsupported_modality",
        "中文解释": "数据类型超出当前 tabular validation 能力",
        "definition": (
            "The artifact belongs to a modality not supported by the current "
            "lightweight validator, such as images, graph databases, binary "
            "scientific formats, MINC, DICOM, VASP, or Neo4j dump files."
        ),
        "example": "Thermal images; .mnc deformation fields; Neo4j dump",
    },
    {
        "category": "inventory_absent",
        "中文解释": "没有可供 grounding 的物理文件清单",
        "definition": (
            "No usable physical candidate files were available for this article, "
            "so grounding could not be attempted."
        ),
        "example": "No physical files available in candidate list",
    },
    {
        "category": "schema_granularity_mismatch",
        "中文解释": "文档逻辑粒度与仓库物理组织粒度不一致",
        "definition": (
            "The logical claim refers to a dataset family, directory, or "
            "abstract group, while the repository realizes it as multiple files, "
            "split files, versions, or directory-level structures."
        ),
        "example": "One logical record set split into ModifiedRecords and RetiredRecords",
    },
    {
        "category": "semantic_ambiguity",
        "中文解释": "存在多个合理候选，无法唯一确定",
        "definition": (
            "Multiple physical artifacts are plausible matches for the same "
            "logical claim, and available evidence is insufficient to choose one."
        ),
        "example": "Two template files both plausibly match the logical template",
    },
]


def build_workbook() -> None:
    if not INPUT_CSV.exists():
        raise FileNotFoundError(f"Input CSV not found: {INPUT_CSV}")

    df = pd.read_csv(INPUT_CSV)

    review = df.copy()

    if "failure_category" in review.columns:
        review = review.rename(
            columns={"failure_category": "auto_category"}
        )

    keep_cols = [
        "article_id",
        "logical_name",
        "status",
        "confidence",
        "reason",
        "auto_category",
    ]

    for col in keep_cols:
        if col not in review.columns:
            review[col] = ""

    review = review[keep_cols]

    review["reviewer_category"] = ""
    review["support_level"] = ""
    review["reviewer_notes"] = ""

    dictionary = pd.DataFrame(CATEGORY_DEFINITIONS)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        review.to_excel(writer, sheet_name="review", index=False)
        dictionary.to_excel(
            writer,
            sheet_name="category_dictionary",
            index=False,
        )

    wb = load_workbook(OUTPUT_XLSX)

    ws = wb["review"]
    dict_ws = wb["category_dictionary"]

    # Header style
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for cell in dict_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    # Freeze panes and filter
    ws.freeze_panes = "A2"
    dict_ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    dict_ws.auto_filter.ref = dict_ws.dimensions

    # Column widths
    widths = {
        "A": 22,  # article_id
        "B": 36,  # logical_name
        "C": 14,  # status
        "D": 12,  # confidence
        "E": 90,  # reason
        "F": 28,  # auto_category
        "G": 32,  # reviewer_category
        "H": 20,  # support_level
        "I": 48,  # reviewer_notes
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for col in range(1, dict_ws.max_column + 1):
        dict_ws.column_dimensions[get_column_letter(col)].width = 35

    # Wrap text
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    for row in dict_ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # Data validation dropdowns
    max_row = max(ws.max_row, 2)

    category_formula = '"' + ",".join(CATEGORY_OPTIONS) + '"'
    support_formula = '"' + ",".join(SUPPORT_OPTIONS) + '"'

    category_dv = DataValidation(
        type="list",
        formula1=category_formula,
        allow_blank=True,
    )
    category_dv.error = "Please choose one of the predefined categories."
    category_dv.errorTitle = "Invalid reviewer_category"

    support_dv = DataValidation(
        type="list",
        formula1=support_formula,
        allow_blank=True,
    )
    support_dv.error = "Please choose one of the predefined support levels."
    support_dv.errorTitle = "Invalid support_level"

    ws.add_data_validation(category_dv)
    ws.add_data_validation(support_dv)

    category_dv.add(f"G2:G{max_row}")
    support_dv.add(f"H2:H{max_row}")

    # Highlight editable columns
    editable_fill = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )

    for row in range(2, max_row + 1):
        ws[f"G{row}"].fill = editable_fill
        ws[f"H{row}"].fill = editable_fill
        ws[f"I{row}"].fill = editable_fill

    # Add dictionary notes sheet styling
    for row in dict_ws.iter_rows(min_row=2):
        dict_ws.row_dimensions[row[0].row].height = 55

    wb.save(OUTPUT_XLSX)

    print(f"Saved: {OUTPUT_XLSX}")
    print(f"Rows for review: {len(review)}")


if __name__ == "__main__":
    build_workbook()
