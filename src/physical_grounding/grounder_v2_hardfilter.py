from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional
import csv
import difflib
import fnmatch
import json
import re


TABULAR_SUFFIXES = {
    ".csv",
    ".tsv",
    ".xlsx",
    ".xls",
    ".parquet",
}

EXCEL_SUFFIXES = {".xlsx", ".xls"}

RESOLVED_THRESHOLD = 0.80
CANDIDATE_THRESHOLD = 0.45
FORMAT_ONLY_SCORE = 0.20

# Only inspect a small number of candidate files per logical claim.
# This keeps grounding fast even when the artifact root contains thousands of files.
MAX_CANDIDATES_TO_OBSERVE = 5
MAX_COLUMN_MATCHES_TO_STORE = 200
MAX_UNMATCHED_COLUMNS_TO_STORE = 50


# ---------------------------------------------------------------------------
# Basic normalization
# ---------------------------------------------------------------------------


def is_unknown(value: Any) -> bool:
    if value is None:
        return True

    text = str(value).strip().lower()

    return text in {
        "",
        "unknown",
        "none",
        "null",
        "n/a",
        "na",
    }


def normalize_text(value: Any) -> str:
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text)
    return text.strip("_")


def split_tokens(value: Any) -> set[str]:
    text = normalize_text(value)

    if not text:
        return set()

    return {
        t for t in text.split("_")
        if t and len(t) >= 2
    }


def safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


# ---------------------------------------------------------------------------
# Format helpers
# ---------------------------------------------------------------------------


def normalize_expected_format(value: Any) -> str:
    fmt = safe_str(value).strip().lower().lstrip(".")

    aliases = {
        "comma separated values": "csv",
        "comma-separated values": "csv",
        "tab separated values": "tsv",
        "tab-separated values": "tsv",
        "excel": "xlsx",
        "excel spreadsheet": "xlsx",
        "microsoft excel": "xlsx",
        "spreadsheet": "xlsx",
        "parquet file": "parquet",
        "text": "txt",
        "plain text": "txt",
        "json file": "json",
    }

    return aliases.get(fmt, fmt)


def get_expected_suffix(expected_format: Any) -> str:
    fmt = normalize_expected_format(expected_format)

    if fmt in {"csv", "tsv", "xlsx", "xls", "parquet", "json", "txt"}:
        return f".{fmt}"

    return ""


def physical_file_candidates(
    physical_inventory: Dict[str, Any],
) -> List[Dict[str, Any]]:
    files = physical_inventory.get("files") or []

    return [
        f for f in files
        if isinstance(f, dict)
    ]


# ---------------------------------------------------------------------------
# Logical claim helpers
# ---------------------------------------------------------------------------


def is_non_file_claim(logical_file: Dict[str, Any]) -> bool:
    """
    Detect logical claims that describe APIs/endpoints/streams rather than
    deposited physical files.

    These should not be treated as ordinary missing physical files.
    """
    text = " ".join(
        str(logical_file.get(k, "") or "")
        for k in [
            "logical_name",
            "documented_path_or_pattern",
            "expected_format",
            "schema_type",
            "role",
        ]
    ).lower()

    non_file_markers = [
        "api",
        "endpoint",
        "restful",
        "rest api",
        "web api",
        "data stream",
        "database query",
        "remote service",
    ]

    return any(marker in text for marker in non_file_markers)


def documented_column_names(logical_file: Dict[str, Any]) -> List[str]:
    names: List[str] = []
    for col in logical_file.get("columns", []) or []:
        if not isinstance(col, dict):
            continue
        name = col.get("name") or col.get("column_name") or col.get("field")
        if not is_unknown(name):
            names.append(str(name).strip())
    return names


def documented_column_records(logical_file: Dict[str, Any]) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []
    for idx, col in enumerate(logical_file.get("columns", []) or [], start=1):
        if not isinstance(col, dict):
            continue
        name = col.get("name") or col.get("column_name") or col.get("field")
        if is_unknown(name):
            continue
        records.append(
            {
                "column_id": col.get("column_id") or f"{logical_file.get('logical_file_id')}_col_{idx:03d}",
                "name": str(name).strip(),
                "data_type": col.get("data_type", "unknown"),
                "semantic_type": col.get("semantic_type", "unknown"),
                "description": col.get("description", "unknown"),
                "unit": col.get("unit"),
                "required": col.get("required", "unknown"),
            }
        )
    return records


# ---------------------------------------------------------------------------
# v1 identity matching: filename/path/suffix
# ---------------------------------------------------------------------------


def score_single_match(
    logical_file: Dict[str, Any],
    physical_file: Dict[str, Any],
) -> Tuple[float, str]:
    """
    Score one logical file claim against one physical file.

    This identity stage only uses filename/path/suffix evidence.
    It does not open files and does not inspect headers.
    """
    pattern = logical_file.get("documented_path_or_pattern", "unknown")
    logical_name = logical_file.get("logical_name", "unknown")
    expected_format = logical_file.get("expected_format", "unknown")

    rel_path = str(physical_file.get("relative_path", "") or "")
    name = str(physical_file.get("name", "") or "")
    stem = str(physical_file.get("stem", "") or "")
    suffix = str(physical_file.get("suffix", "") or "").lower()

    pattern_str = "" if is_unknown(pattern) else str(pattern).strip()
    pattern_name = Path(pattern_str).name if pattern_str else ""
    pattern_stem = Path(pattern_name).stem if pattern_name else ""

    expected_suffix = get_expected_suffix(expected_format)

    # 1. Exact relative path.
    if pattern_str and rel_path == pattern_str:
        return 1.00, "exact_relative_path"

    # 2. Exact basename.
    if pattern_name and name == pattern_name:
        return 0.95, "exact_basename"

    # 3. Glob pattern match.
    if pattern_str and any(ch in pattern_str for ch in ["*", "?", "["]):
        if fnmatch.fnmatch(rel_path, pattern_str) or fnmatch.fnmatch(name, pattern_str):
            return 0.90, "glob_pattern"

    # 4. Case-insensitive stem match.
    if pattern_stem and normalize_text(pattern_stem) == normalize_text(stem):
        return 0.85, "stem_match"

    # 5. Normalized basename match.
    if pattern_name and normalize_text(pattern_name) == normalize_text(name):
        return 0.80, "normalized_basename"

    # 6. One stem contains the other.
    norm_pattern_stem = normalize_text(pattern_stem)
    norm_stem = normalize_text(stem)

    if norm_pattern_stem and norm_stem:
        if norm_pattern_stem in norm_stem or norm_stem in norm_pattern_stem:
            return 0.65, "stem_contains"

    # 7. Token overlap between logical name / documented pattern and physical name.
    logical_tokens = split_tokens(logical_name) | split_tokens(pattern_stem)
    physical_tokens = split_tokens(stem)

    if logical_tokens and physical_tokens:
        overlap = logical_tokens & physical_tokens
        union = logical_tokens | physical_tokens
        jaccard = len(overlap) / len(union)

        if jaccard >= 0.50:
            return 0.60, "token_overlap_high"

        if jaccard >= 0.25:
            return 0.45, "token_overlap_low"

    # 8. Format-only match is too weak for identity grounding, but useful as evidence.
    if expected_suffix and suffix == expected_suffix:
        return FORMAT_ONLY_SCORE, "format_only"

    return 0.0, "no_match"


def score_logical_file_against_inventory(
    logical_file: Dict[str, Any],
    physical_files: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    scored = []

    for physical_file in physical_files:
        score, match_type = score_single_match(
            logical_file=logical_file,
            physical_file=physical_file,
        )

        if score <= 0:
            continue

        scored.append(
            {
                "relative_path": physical_file.get("relative_path"),
                "absolute_path": physical_file.get("absolute_path"),
                "name": physical_file.get("name"),
                "stem": physical_file.get("stem"),
                "suffix": physical_file.get("suffix"),
                "size_bytes": physical_file.get("size_bytes"),
                "is_tabular_candidate": physical_file.get("is_tabular_candidate"),
                "match_score": round(score, 4),
                "match_type": match_type,
            }
        )

    scored.sort(
        key=lambda x: x["match_score"],
        reverse=True,
    )

    return scored


def determine_grounding_status(
    logical_file: Dict[str, Any],
    candidates: List[Dict[str, Any]],
) -> Tuple[str, List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Determine v1 identity grounding status.

    Returns:
      status,
      matched_physical_files,
      candidate_physical_files,
      warnings
    """
    warnings = []

    logical_name = logical_file.get("logical_name", "unknown")
    pattern = logical_file.get("documented_path_or_pattern", "unknown")

    if is_non_file_claim(logical_file):
        warnings.append(
            {
                "warning_type": "unsupported_non_file_claim",
                "severity": "medium",
                "message": (
                    "This logical claim appears to describe an API, endpoint, "
                    "or data stream rather than a deposited physical file. "
                    "Physical file grounding cannot resolve it."
                ),
            }
        )

        return "unsupported_non_file_claim", [], [], warnings

    if is_unknown(logical_name) and is_unknown(pattern):
        warnings.append(
            {
                "warning_type": "ungroundable_logical_claim",
                "severity": "blocking",
                "message": (
                    "Logical file has neither logical_name nor "
                    "documented_path_or_pattern."
                ),
            }
        )

        return "ungroundable", [], [], warnings

    if not candidates:
        warnings.append(
            {
                "warning_type": "missing_physical_match",
                "severity": "high",
                "message": (
                    "No physical artifact candidate matched this logical file claim."
                ),
            }
        )

        return "missing", [], [], warnings

    top_score = candidates[0]["match_score"]
    candidate_files = candidates[:10]

    strong_candidates = [
        c for c in candidates
        if c["match_score"] >= CANDIDATE_THRESHOLD
    ]

    if not strong_candidates:
        format_only_candidates = [
            c for c in candidates
            if c["match_type"] == "format_only"
        ]

        if format_only_candidates:
            warnings.append(
                {
                    "warning_type": "missing_name_match_with_same_format_candidates",
                    "severity": "high",
                    "message": (
                        "Physical files with the expected format exist, but none "
                        "match the documented file name or pattern."
                    ),
                }
            )
        else:
            warnings.append(
                {
                    "warning_type": "missing_physical_match",
                    "severity": "high",
                    "message": (
                        "No physical artifact candidate matched this logical file "
                        "claim."
                    ),
                }
            )

        return "missing", [], candidate_files, warnings

    close_candidates = [
        c for c in strong_candidates
        if c["match_score"] >= top_score - 0.05
    ]

    if top_score >= RESOLVED_THRESHOLD and len(close_candidates) == 1:
        matched = [candidates[0]]
        return "resolved", matched, candidate_files, warnings

    if len(close_candidates) > 1:
        warnings.append(
            {
                "warning_type": "ambiguous_physical_match",
                "severity": "medium",
                "message": (
                    "Multiple physical artifacts match the logical file claim "
                    "with similar scores."
                ),
            }
        )

        return "ambiguous", close_candidates, candidate_files, warnings

    if top_score >= CANDIDATE_THRESHOLD:
        warnings.append(
            {
                "warning_type": "weak_physical_match",
                "severity": "medium",
                "message": (
                    "A possible physical artifact match was found, but the match "
                    "is weak and should be reviewed by a curator."
                ),
            }
        )

        return "weak_match", [candidates[0]], candidate_files, warnings

    warnings.append(
        {
            "warning_type": "missing_physical_match",
            "severity": "high",
            "message": (
                "No sufficiently strong physical artifact match was found."
            ),
        }
    )

    return "missing", [], candidate_files, warnings


# ---------------------------------------------------------------------------
# Physical observation and header extraction
# ---------------------------------------------------------------------------


def read_csv_like_header(path: Path, suffix: str) -> Dict[str, Any]:
    delimiter = "\t" if suffix == ".tsv" else None
    encodings = ["utf-8-sig", "utf-8", "latin-1", "cp1252"]

    last_error = ""
    for enc in encodings:
        try:
            with path.open("r", encoding=enc, errors="strict", newline="") as f:
                sample = f.read(65536)
                if not sample:
                    return {
                        "load_status": "empty_file",
                        "loader": "csv",
                        "encoding": enc,
                        "delimiter": delimiter or ",",
                        "columns": [],
                        "num_columns": 0,
                    }
                f.seek(0)
                if delimiter is None:
                    try:
                        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
                        delimiter_use = dialect.delimiter
                    except Exception:
                        delimiter_use = ","
                else:
                    delimiter_use = delimiter
                reader = csv.reader(f, delimiter=delimiter_use)
                header = next(reader, [])
                header = [str(x).strip() for x in header if str(x).strip()]
                return {
                    "load_status": "success",
                    "loader": "csv",
                    "encoding": enc,
                    "delimiter": delimiter_use,
                    "columns": header,
                    "num_columns": len(header),
                }
        except Exception as exc:  # noqa: BLE001
            last_error = f"{exc.__class__.__name__}: {exc}"

    return {
        "load_status": "header_read_failed",
        "loader": "csv",
        "error": last_error,
        "columns": [],
        "num_columns": 0,
    }


def read_excel_header(path: Path, suffix: str) -> Dict[str, Any]:
    if suffix == ".xls":
        return {
            "load_status": "unsupported_xls_header_without_xlrd",
            "loader": "excel",
            "columns": [],
            "num_columns": 0,
            "note": "Install xlrd or convert .xls files if xls header extraction is required.",
        }

    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # noqa: BLE001
        return {
            "load_status": "missing_optional_dependency_openpyxl",
            "loader": "openpyxl",
            "error": str(exc),
            "columns": [],
            "num_columns": 0,
        }

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        if not sheet_names:
            return {
                "load_status": "no_sheets",
                "loader": "openpyxl",
                "columns": [],
                "num_columns": 0,
            }
        ws = wb[sheet_names[0]]
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
        header = [str(x).strip() for x in first_row if x is not None and str(x).strip()]
        return {
            "load_status": "success",
            "loader": "openpyxl",
            "sheet_name": sheet_names[0],
            "sheet_names": sheet_names[:20],
            "columns": header,
            "num_columns": len(header),
        }
    except Exception as exc:  # noqa: BLE001
        return {
            "load_status": "header_read_failed",
            "loader": "openpyxl",
            "error": f"{exc.__class__.__name__}: {exc}",
            "columns": [],
            "num_columns": 0,
        }


def read_parquet_header(path: Path) -> Dict[str, Any]:
    try:
        import pandas as pd  # type: ignore
        cols = list(pd.read_parquet(path).head(0).columns)
        return {
            "load_status": "success",
            "loader": "pandas_parquet",
            "columns": [str(c) for c in cols],
            "num_columns": len(cols),
        }
    except Exception as exc:  # noqa: BLE001
        return {
            "load_status": "header_read_failed",
            "loader": "pandas_parquet",
            "error": f"{exc.__class__.__name__}: {exc}",
            "columns": [],
            "num_columns": 0,
        }


def observe_physical_file(candidate: Dict[str, Any]) -> Dict[str, Any]:
    path_text = candidate.get("absolute_path") or ""
    suffix = str(candidate.get("suffix") or Path(str(candidate.get("relative_path") or "")).suffix).lower()
    obs = {
        "relative_path": candidate.get("relative_path"),
        "absolute_path": path_text,
        "suffix": suffix,
        "size_bytes": candidate.get("size_bytes"),
        "is_tabular_candidate": suffix in TABULAR_SUFFIXES,
        "header_observation": {
            "load_status": "not_attempted",
            "columns": [],
            "num_columns": 0,
        },
    }

    path = Path(path_text) if path_text else None
    if not path or not path.exists() or not path.is_file():
        obs["header_observation"] = {
            "load_status": "missing_physical_file",
            "columns": [],
            "num_columns": 0,
        }
        return obs

    if suffix in {".csv", ".tsv"}:
        obs["header_observation"] = read_csv_like_header(path, suffix)
    elif suffix in EXCEL_SUFFIXES:
        obs["header_observation"] = read_excel_header(path, suffix)
    elif suffix == ".parquet":
        obs["header_observation"] = read_parquet_header(path)
    else:
        obs["header_observation"] = {
            "load_status": "not_tabular_or_unsupported_for_header",
            "columns": [],
            "num_columns": 0,
        }

    return obs


# ---------------------------------------------------------------------------
# v2 presentation grounding dimensions
# ---------------------------------------------------------------------------


def build_identity_grounding(status: str, matched_files: List[Dict[str, Any]], candidate_files: List[Dict[str, Any]]) -> Dict[str, Any]:
    best = None
    if matched_files:
        best = matched_files[0]
    elif candidate_files:
        best = candidate_files[0]

    if best:
        score = float(best.get("match_score") or 0.0)
        evidence_type = best.get("match_type", "unknown")
    else:
        score = 0.0
        evidence_type = "no_candidate"

    return {
        "dimension": "identity",
        "status": status,
        "score": round(score, 4),
        "evidence_type": evidence_type,
        "best_candidate_path": best.get("relative_path") if best else "",
    }


def ground_format(logical_file: Dict[str, Any], observation: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    expected = normalize_expected_format(logical_file.get("expected_format"))
    expected_suffix = get_expected_suffix(expected)

    if not observation:
        return {
            "dimension": "format",
            "status": "no_physical_candidate",
            "score": 0.0,
            "expected_format": expected or "unknown",
            "observed_suffix": "",
        }

    observed_suffix = str(observation.get("suffix") or "").lower()
    header_status = (observation.get("header_observation") or {}).get("load_status", "not_attempted")

    if is_unknown(expected):
        status = "expected_format_unknown"
        score = 0.5 if observed_suffix in TABULAR_SUFFIXES else 0.25
    elif expected_suffix and observed_suffix == expected_suffix:
        status = "supported"
        score = 1.0
    elif expected in {"xlsx", "xls", "excel", "spreadsheet"} and observed_suffix in EXCEL_SUFFIXES:
        status = "compatible_excel_family"
        score = 0.9
    elif expected in {"csv", "tsv", "xlsx", "xls", "parquet", "tabular"} and observed_suffix in TABULAR_SUFFIXES:
        status = "tabular_family_mismatch"
        score = 0.65
    else:
        status = "mismatch"
        score = 0.0

    if header_status in {"success", "empty_file"} and score > 0:
        loader_support = "loadable_or_header_readable"
    elif header_status == "not_tabular_or_unsupported_for_header":
        loader_support = "not_attempted_for_non_tabular"
    else:
        loader_support = header_status

    return {
        "dimension": "format",
        "status": status,
        "score": round(score, 4),
        "expected_format": expected or "unknown",
        "expected_suffix": expected_suffix,
        "observed_suffix": observed_suffix,
        "header_load_status": header_status,
        "loader_support": loader_support,
    }


def similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    if normalize_text(a) == normalize_text(b):
        return 1.0
    return difflib.SequenceMatcher(None, normalize_text(a), normalize_text(b)).ratio()


def match_one_column(doc_col: Dict[str, Any], observed_columns: List[str]) -> Dict[str, Any]:
    name = safe_str(doc_col.get("name"))
    semantic_type = safe_str(doc_col.get("semantic_type"))
    description = safe_str(doc_col.get("description"))

    if not observed_columns:
        return {
            "documented_column": name,
            "matched_column": "",
            "match_type": "no_observed_columns",
            "score": 0.0,
        }

    norm_name = normalize_text(name)
    observed_norm = {normalize_text(c): c for c in observed_columns}

    if norm_name in observed_norm:
        return {
            "documented_column": name,
            "matched_column": observed_norm[norm_name],
            "match_type": "normalized_exact",
            "score": 1.0,
        }

    best_col = ""
    best_score = 0.0
    for obs in observed_columns:
        s = similarity(name, obs)
        if s > best_score:
            best_score = s
            best_col = obs

    if best_score >= 0.86:
        return {
            "documented_column": name,
            "matched_column": best_col,
            "match_type": "fuzzy_name_high",
            "score": round(best_score, 4),
        }

    # Weak semantic/name evidence from semantic_type or description tokens.
    doc_tokens = split_tokens(name) | split_tokens(semantic_type) | split_tokens(description)
    best_sem_col = ""
    best_sem_score = 0.0
    for obs in observed_columns:
        obs_tokens = split_tokens(obs)
        if not doc_tokens or not obs_tokens:
            continue
        overlap = doc_tokens & obs_tokens
        union = doc_tokens | obs_tokens
        score = len(overlap) / len(union) if union else 0.0
        if score > best_sem_score:
            best_sem_score = score
            best_sem_col = obs

    if best_sem_score >= 0.30:
        return {
            "documented_column": name,
            "matched_column": best_sem_col,
            "match_type": "semantic_token_overlap",
            "score": round(best_sem_score, 4),
        }

    if best_score >= 0.65:
        return {
            "documented_column": name,
            "matched_column": best_col,
            "match_type": "fuzzy_name_low",
            "score": round(best_score, 4),
        }

    return {
        "documented_column": name,
        "matched_column": "",
        "match_type": "unmatched",
        "score": round(max(best_score, best_sem_score), 4),
    }


def ground_columns(
    logical_file: Dict[str, Any],
    observation: Optional[Dict[str, Any]],
    file_confirmed: bool,
) -> Dict[str, Any]:
    doc_cols = documented_column_records(logical_file)
    num_doc = len(doc_cols)

    if num_doc == 0:
        return {
            "dimension": "columns",
            "status": "no_documented_columns",
            "score": None,
            "num_documented_columns": 0,
            "num_observed_columns": 0,
            "num_matched_columns": 0,
            "coverage": None,
            "matches": [],
            "unmatched_documented_columns": [],
        }

    if not file_confirmed:
        return {
            "dimension": "columns",
            "status": "not_applicable_file_not_confirmed",
            "score": None,
            "num_documented_columns": num_doc,
            "num_observed_columns": 0,
            "num_matched_columns": 0,
            "coverage": None,
            "matches": [],
            "unmatched_documented_columns": [],
            "note": (
                "Column-wise grounding is only attempted when the logical file claim "
                "has a machine-confirmed physical file match. Review identity grounding first."
            ),
        }

    if not observation:
        return {
            "dimension": "columns",
            "status": "no_physical_candidate",
            "score": 0.0,
            "num_documented_columns": num_doc,
            "num_observed_columns": 0,
            "num_matched_columns": 0,
            "coverage": 0.0,
            "matches": [],
            "unmatched_documented_columns": [c["name"] for c in doc_cols[:MAX_UNMATCHED_COLUMNS_TO_STORE]],
        }

    header_obs = observation.get("header_observation") or {}
    observed_columns = [str(c) for c in header_obs.get("columns", []) or [] if str(c).strip()]

    if not observed_columns:
        return {
            "dimension": "columns",
            "status": "no_observed_headers",
            "score": 0.0,
            "num_documented_columns": num_doc,
            "num_observed_columns": 0,
            "num_matched_columns": 0,
            "coverage": 0.0,
            "header_load_status": header_obs.get("load_status", "unknown"),
            "matches": [],
            "unmatched_documented_columns": [c["name"] for c in doc_cols[:MAX_UNMATCHED_COLUMNS_TO_STORE]],
        }

    matches = [match_one_column(c, observed_columns) for c in doc_cols]
    matched = [m for m in matches if float(m.get("score") or 0.0) >= 0.65]
    strong = [m for m in matches if float(m.get("score") or 0.0) >= 0.86]
    coverage = len(matched) / num_doc if num_doc else 0.0
    strong_coverage = len(strong) / num_doc if num_doc else 0.0

    if coverage >= 0.80:
        status = "supported"
    elif coverage >= 0.50:
        status = "partially_supported"
    elif coverage > 0:
        status = "weakly_supported"
    else:
        status = "not_supported"

    unmatched = [m["documented_column"] for m in matches if float(m.get("score") or 0.0) < 0.65]

    return {
        "dimension": "columns",
        "status": status,
        "score": round(coverage, 4),
        "num_documented_columns": num_doc,
        "num_observed_columns": len(observed_columns),
        "num_matched_columns": len(matched),
        "num_strong_matched_columns": len(strong),
        "coverage": round(coverage, 4),
        "strong_coverage": round(strong_coverage, 4),
        "header_load_status": header_obs.get("load_status", "unknown"),
        "observed_columns_sample": observed_columns[:30],
        "matches": matches[:MAX_COLUMN_MATCHES_TO_STORE],
        "unmatched_documented_columns": unmatched[:MAX_UNMATCHED_COLUMNS_TO_STORE],
    }


def ground_role(logical_file: Dict[str, Any], observation: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    role = logical_file.get("role", "unknown")
    schema_type = logical_file.get("schema_type", "unknown")
    logical_name = logical_file.get("logical_name", "unknown")

    if is_unknown(role) and is_unknown(schema_type):
        return {
            "dimension": "role",
            "status": "role_unknown",
            "score": None,
            "expected_role": role,
            "schema_type": schema_type,
        }

    logical_tokens = split_tokens(role) | split_tokens(schema_type) | split_tokens(logical_name)

    if not observation:
        return {
            "dimension": "role",
            "status": "no_physical_candidate",
            "score": 0.0,
            "expected_role": role,
            "schema_type": schema_type,
        }

    header_obs = observation.get("header_observation") or {}
    observed_columns = header_obs.get("columns", []) or []
    physical_text = " ".join(
        [
            safe_str(observation.get("relative_path")),
            safe_str(Path(safe_str(observation.get("relative_path"))).stem),
            " ".join(safe_str(c) for c in observed_columns[:50]),
        ]
    )
    physical_tokens = split_tokens(physical_text)

    if not logical_tokens or not physical_tokens:
        return {
            "dimension": "role",
            "status": "insufficient_role_evidence",
            "score": 0.0,
            "expected_role": role,
            "schema_type": schema_type,
        }

    overlap = logical_tokens & physical_tokens
    union = logical_tokens | physical_tokens
    score = len(overlap) / len(union) if union else 0.0

    if score >= 0.25:
        status = "supported"
    elif score > 0:
        status = "weakly_supported"
    else:
        status = "not_supported"

    return {
        "dimension": "role",
        "status": status,
        "score": round(score, 4),
        "expected_role": role,
        "schema_type": schema_type,
        "overlap_tokens": sorted(overlap)[:20],
    }


def summarize_presentation_grounding(
    identity_grounding: Dict[str, Any],
    format_grounding: Dict[str, Any],
    role_grounding: Dict[str, Any],
    column_grounding: Dict[str, Any],
) -> Dict[str, Any]:
    components: List[Tuple[str, float, float]] = []

    components.append(("identity", float(identity_grounding.get("score") or 0.0), 0.40))

    if format_grounding.get("score") is not None:
        components.append(("format", float(format_grounding.get("score") or 0.0), 0.20))

    if role_grounding.get("score") is not None:
        components.append(("role", float(role_grounding.get("score") or 0.0), 0.10))

    if column_grounding.get("score") is not None:
        components.append(("columns", float(column_grounding.get("score") or 0.0), 0.30))

    if not components:
        score = 0.0
    else:
        denom = sum(w for _, _, w in components)
        score = sum(s * w for _, s, w in components) / denom if denom else 0.0

    failed_dimensions: List[str] = []
    uncertain_dimensions: List[str] = []

    identity_status = identity_grounding.get("status")
    if identity_status not in {"resolved"}:
        failed_dimensions.append("identity")

    if format_grounding.get("status") in {"mismatch", "no_physical_candidate"}:
        failed_dimensions.append("format")
    elif format_grounding.get("status") in {"expected_format_unknown", "tabular_family_mismatch"}:
        uncertain_dimensions.append("format")

    if role_grounding.get("status") in {"not_supported", "no_physical_candidate"}:
        failed_dimensions.append("role")
    elif role_grounding.get("status") in {"role_unknown", "insufficient_role_evidence", "weakly_supported"}:
        uncertain_dimensions.append("role")

    if column_grounding.get("status") in {"no_physical_candidate", "no_observed_headers", "not_supported"}:
        failed_dimensions.append("columns")
    elif column_grounding.get("status") in {"partially_supported", "weakly_supported"}:
        uncertain_dimensions.append("columns")
    elif column_grounding.get("status") == "not_applicable_file_not_confirmed":
        # Do not create a separate column-level warning when identity grounding is unresolved.
        # The file-level identity issue is the annotation target.
        pass

    return {
        "score": round(score, 4),
        "components": [
            {"dimension": d, "score": round(s, 4), "weight": w}
            for d, s, w in components
        ],
        "failed_dimensions": sorted(set(failed_dimensions)),
        "uncertain_dimensions": sorted(set(uncertain_dimensions)),
    }


def generate_annotation_target(
    logical_file: Dict[str, Any],
    status: str,
    warnings: List[Dict[str, Any]],
    presentation_grounding: Dict[str, Any],
) -> Dict[str, Any]:
    """
    File-level annotation target.

    Important boundary:
      file-level annotation is only for unresolved file identity/mapping.
      resolved files should not enter the file annotation table merely because
      role evidence is weak or column grounding has uncertainty. Column issues
      are exported separately as column-wise annotation targets.
    """
    needs_file_identity_review = status != "resolved"

    if status == "resolved":
        failed = []
        uncertain = []
    else:
        failed = ["identity"]
        uncertain = []

    question = (
        "Is this file-level grounding disagreement caused by a documentation-artifact mismatch, "
        "an LLM extraction error, a grounding algorithm error, insufficient documentation, "
        "or an acceptable variant/manual mapping?"
    )

    return {
        "needs_review": needs_file_identity_review,
        "annotation_scope": "file_identity_grounding",
        "failed_dimensions": failed,
        "uncertain_dimensions": uncertain,
        "suggested_human_question": question,
        "human_error_source_options": [
            "documentation_artifact_mismatch",
            "llm_extraction_error",
            "grounding_algorithm_error",
            "insufficient_documentation",
            "acceptable_variant",
            "unsupported_or_non_file_claim",
            "artifact_access_or_unpacking_issue",
            "other_uncertain",
        ],
        "human_action_options": [
            "accept_system_assessment",
            "override_grounding",
            "manual_file_mapping",
            "mark_missing",
            "mark_ambiguous",
            "accept_variant_mapping",
            "exclude_from_file_validation",
            "request_supplementary_materials",
            "needs_second_review",
        ],
    }


# ---------------------------------------------------------------------------
# Main file grounding
# ---------------------------------------------------------------------------


def ground_one_logical_file(
    logical_file: Dict[str, Any],
    physical_files: List[Dict[str, Any]],
) -> Dict[str, Any]:
    candidates = score_logical_file_against_inventory(
        logical_file=logical_file,
        physical_files=physical_files,
    )

    status, matched_files, candidate_files, warnings = determine_grounding_status(
        logical_file=logical_file,
        candidates=candidates,
    )

    # Observe only matched files if available, otherwise top candidate files.
    observation_candidates = (matched_files or candidate_files)[:MAX_CANDIDATES_TO_OBSERVE]
    observed_candidates = []
    for c in observation_candidates:
        c2 = dict(c)
        c2["physical_observation"] = observe_physical_file(c)
        observed_candidates.append(c2)

    best_observation = None
    if observed_candidates:
        best_observation = observed_candidates[0].get("physical_observation")

    identity_grounding = build_identity_grounding(status, matched_files, candidate_files)
    format_grounding = ground_format(logical_file, best_observation)
    role_grounding = ground_role(logical_file, best_observation)
    # Column-wise grounding is only valid when the machine has confirmed the file identity.
    # Ambiguous/weak/missing file matches remain file-level annotation targets.
    file_confirmed_for_columns = status == "resolved" and bool(matched_files)
    column_grounding = ground_columns(
        logical_file,
        best_observation if file_confirmed_for_columns else None,
        file_confirmed=file_confirmed_for_columns,
    )

    presentation_grounding = summarize_presentation_grounding(
        identity_grounding=identity_grounding,
        format_grounding=format_grounding,
        role_grounding=role_grounding,
        column_grounding=column_grounding,
    )

    annotation_target = generate_annotation_target(
        logical_file=logical_file,
        status=status,
        warnings=warnings,
        presentation_grounding=presentation_grounding,
    )

    human_review_needed = bool(annotation_target.get("needs_review"))

    return {
        "logical_file_id": logical_file.get("logical_file_id"),
        "logical_name": logical_file.get("logical_name"),
        "documented_path_or_pattern": logical_file.get("documented_path_or_pattern"),
        "expected_format": logical_file.get("expected_format"),
        "schema_type": logical_file.get("schema_type"),
        "role": logical_file.get("role"),
        "num_documented_columns": len(logical_file.get("columns", []) or []),

        # v1 identity status retained for backward compatibility.
        "grounding_status": status,
        "human_review_needed": human_review_needed,

        "matched_physical_files": observed_candidates if matched_files else [],
        "candidate_physical_files": observed_candidates if not matched_files else candidate_files,
        "warnings": warnings,

        # v2 multi-dimensional presentation grounding.
        "identity_grounding": identity_grounding,
        "format_grounding": format_grounding,
        "role_grounding": role_grounding,
        "column_grounding": column_grounding,
        "presentation_grounding": presentation_grounding,
        "annotation_target": annotation_target,
    }


# ---------------------------------------------------------------------------
# Summary and review target generation
# ---------------------------------------------------------------------------


def summarize_file_groundings(
    file_groundings: List[Dict[str, Any]],
) -> Dict[str, Any]:
    total = len(file_groundings)

    counts = {
        "resolved": 0,
        "ambiguous": 0,
        "weak_match": 0,
        "missing": 0,
        "ungroundable": 0,
        "unsupported_non_file_claim": 0,
    }

    for g in file_groundings:
        status = g.get("grounding_status", "ungroundable")

        if status not in counts:
            counts[status] = 0

        counts[status] += 1

    file_applicable_total = (
        total
        - counts.get("unsupported_non_file_claim", 0)
    )

    if file_applicable_total <= 0:
        grounding_score = 0.0
    else:
        grounding_score = (
            counts.get("resolved", 0)
            + 0.50 * counts.get("ambiguous", 0)
            + 0.25 * counts.get("weak_match", 0)
        ) / file_applicable_total

    warning_counts: Dict[str, int] = {}

    for g in file_groundings:
        for w in g.get("warnings", []):
            wtype = w.get("warning_type", "unknown")
            warning_counts[wtype] = warning_counts.get(wtype, 0) + 1

    # v2 summaries.
    dim_status_counts: Dict[str, Dict[str, int]] = {
        "identity": {},
        "format": {},
        "role": {},
        "columns": {},
    }
    presentation_scores: List[float] = []
    column_coverages: List[float] = []
    total_documented_columns = 0
    total_matched_columns = 0
    num_column_grounding_applicable = 0

    for g in file_groundings:
        for dim, key in [
            ("identity", "identity_grounding"),
            ("format", "format_grounding"),
            ("role", "role_grounding"),
            ("columns", "column_grounding"),
        ]:
            status_value = (g.get(key) or {}).get("status", "missing")
            dim_status_counts[dim][status_value] = dim_status_counts[dim].get(status_value, 0) + 1

        pg = g.get("presentation_grounding") or {}
        if pg.get("score") is not None:
            presentation_scores.append(float(pg.get("score") or 0.0))

        cg = g.get("column_grounding") or {}
        # Only confirmed physical-file matches enter column-wise grounding workload/statistics.
        if cg.get("coverage") is not None:
            num_column_grounding_applicable += 1
            total_documented_columns += int(cg.get("num_documented_columns") or 0)
            total_matched_columns += int(cg.get("num_matched_columns") or 0)
            column_coverages.append(float(cg.get("coverage") or 0.0))

    mean_presentation_score = (
        sum(presentation_scores) / len(presentation_scores)
        if presentation_scores else 0.0
    )
    mean_column_coverage = (
        sum(column_coverages) / len(column_coverages)
        if column_coverages else 0.0
    )
    overall_column_coverage = (
        total_matched_columns / total_documented_columns
        if total_documented_columns else None
    )

    return {
        "num_logical_files": total,
        "num_file_grounding_applicable": file_applicable_total,
        "num_resolved": counts.get("resolved", 0),
        "num_ambiguous": counts.get("ambiguous", 0),
        "num_weak_match": counts.get("weak_match", 0),
        "num_missing": counts.get("missing", 0),
        "num_ungroundable": counts.get("ungroundable", 0),
        "num_unsupported_non_file_claim": counts.get(
            "unsupported_non_file_claim",
            0,
        ),
        "num_human_review_needed": sum(
            1 for g in file_groundings
            if g.get("human_review_needed")
        ),
        "grounding_score": round(grounding_score, 4),
        "warning_counts": warning_counts,

        # v2 aggregate presentation grounding statistics.
        "presentation_grounding_score": round(mean_presentation_score, 4),
        "dimension_status_counts": dim_status_counts,
        "num_column_grounding_applicable": num_column_grounding_applicable,
        "total_documented_columns": total_documented_columns,
        "total_matched_columns": total_matched_columns,
        "mean_file_column_coverage": round(mean_column_coverage, 4),
        "overall_column_coverage": round(overall_column_coverage, 4) if overall_column_coverage is not None else None,
    }


def generate_human_review_targets(
    file_groundings: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    targets = []

    for g in file_groundings:
        annotation_target = g.get("annotation_target") or {}
        if not annotation_target.get("needs_review"):
            continue

        targets.append(
            {
                "target_id": f"review_{g.get('logical_file_id')}",
                "logical_file_id": g.get("logical_file_id"),
                "logical_name": g.get("logical_name"),
                "grounding_status": g.get("grounding_status"),
                "failed_dimensions": annotation_target.get("failed_dimensions", []),
                "uncertain_dimensions": annotation_target.get("uncertain_dimensions", []),
                "question": annotation_target.get("suggested_human_question"),
                "default_action": "needs_second_review",
            }
        )

    return targets


def generate_global_warnings(
    grounding_inputs: Dict[str, Any],
) -> List[Dict[str, Any]]:
    warnings = []

    physical_inventory = grounding_inputs.get("physical_inventory", {})
    artifact_sources_overview = grounding_inputs.get("artifact_sources_overview", {})

    if not physical_inventory.get("effective_artifact_root_exists", False):
        warnings.append(
            {
                "warning_type": "missing_effective_artifact_root",
                "severity": "blocking",
                "message": "The effective artifact root does not exist.",
            }
        )

    if physical_inventory.get("num_files", 0) == 0:
        warnings.append(
            {
                "warning_type": "empty_effective_artifact_root",
                "severity": "blocking",
                "message": "The effective artifact root contains no files.",
            }
        )

    existing_sources = [
        name for name, info in artifact_sources_overview.items()
        if info.get("exists") and info.get("num_files", 0) > 0
    ]

    if len(existing_sources) > 1:
        warnings.append(
            {
                "warning_type": "multiple_artifact_sources_available",
                "severity": "low",
                "message": (
                    "Multiple artifact sources are available. Grounding uses the "
                    "effective artifact root only, but other sources may contain "
                    "additional or duplicate files."
                ),
                "sources": existing_sources,
            }
        )

    return warnings


class RefinedArtifactGrounder:
    """
    Rule-based physical grounding for selected-50.

    Input:
      logical claims from dataset_structure.json
      physical inventory from effective_artifact_root

    Output:
      refined_artifact_grounding.json

    This v2 version retains v1 filename/path identity grounding and adds
    multi-dimensional presentation grounding:
      - identity grounding
      - format grounding
      - role grounding
      - column/header grounding
      - annotation target generation
    """

    def ground(
        self,
        grounding_inputs: Dict[str, Any],
    ) -> Dict[str, Any]:
        article_id = grounding_inputs.get("article_id", "unknown")

        logical_claims = grounding_inputs.get("logical_claims", {})
        physical_inventory = grounding_inputs.get("physical_inventory", {})

        logical_files = logical_claims.get("logical_files") or []
        physical_files = physical_file_candidates(physical_inventory)

        file_groundings = [
            ground_one_logical_file(
                logical_file=logical_file,
                physical_files=physical_files,
            )
            for logical_file in logical_files
        ]

        summary = summarize_file_groundings(file_groundings)
        global_warnings = generate_global_warnings(grounding_inputs)
        human_review_targets = generate_human_review_targets(file_groundings)

        return {
            "schema_version": "refined_artifact_grounding_v2",
            "generated_at": datetime.now().isoformat(),
            "article_id": article_id,
            "input_layers": {
                "logical_schema_path": logical_claims.get("schema_path"),
                "effective_artifact_root": physical_inventory.get(
                    "effective_artifact_root"
                ),
                "effective_artifact_source": physical_inventory.get(
                    "effective_artifact_source"
                ),
            },
            "logical_summary": logical_claims.get("summary", {}),
            "physical_summary": {
                "num_physical_files": physical_inventory.get("num_files"),
                "num_tabular_candidates": physical_inventory.get(
                    "num_tabular_candidates"
                ),
                "effective_artifact_root_exists": physical_inventory.get(
                    "effective_artifact_root_exists"
                ),
            },
            "summary": summary,
            "file_groundings": file_groundings,
            "human_review_targets": human_review_targets,
            "global_warnings": global_warnings,
        }
